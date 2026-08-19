import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def clean(value):
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"[\u2066-\u2069\u200e\u200f]", "", text)
    return re.sub(r"[ \t]+", " ", text).strip()


def sheet_month_year(title):
    lower = title.strip().lower()
    match = re.match(r"([a-z]+)\s+(\d{4})$", lower)
    if match and match.group(1) in MONTHS:
        return MONTHS[match.group(1)], int(match.group(2))
    match = re.match(r"(\d{2})\.(\d{4})$", lower)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None


def parse_time(token):
    token = token.strip().upper().replace(".", "").replace(" ", "")
    token = token.replace("NN", "PM")
    if token in {"MORNING"}:
        return "09:00"
    if token in {"AFTERNOON"}:
        return "13:00"
    if token in {"EVENING"}:
        return "18:00"
    if token in {"ONWARDS"}:
        return None
    match = re.match(r"(\d{1,2})(?::(\d{2}))?(AM|PM)?$", token)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or "0")
    suffix = match.group(3)
    if suffix == "PM" and hour != 12:
        hour += 12
    if suffix == "AM" and hour == 12:
        hour = 0
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return f"{hour:02d}:{minute:02d}"
    return None


def date_range_and_time(value, sheet_month, sheet_year):
    if isinstance(value, datetime):
        start = value
        return start.date(), start.date(), "09:00", "10:00"
    text = clean(value)
    if not text:
        return None
    lines = [line.strip() for line in re.split(r"\n+", text) if line.strip()]
    first = lines[0] if lines else text
    if re.match(r"\d{4}-\d{2}-\d{2}", first):
        dt = datetime.fromisoformat(first.split()[0])
        return dt.date(), dt.date(), "09:00", "10:00"

    month_names = "|".join(sorted(MONTHS, key=len, reverse=True))
    match = re.search(rf"({month_names})\s+(\d{{1,2}})(?:\s*(?:to|-|and|or)\s*(?:(?:{month_names})\s+)?(\d{{1,2}}))?", first, re.I)
    if not match:
        return None
    month = MONTHS[match.group(1).lower()]
    day1 = int(match.group(2))
    day2 = int(match.group(3) or day1)
    start_date = datetime(sheet_year, month, day1).date()
    end_date = datetime(sheet_year, month, day2).date()
    if end_date < start_date:
        end_date = start_date

    time_text = " ".join(lines[2:] if len(lines) > 2 else lines[1:])
    time_text = time_text.replace("–", " to ").replace("-", " to ")
    time_text = re.sub(r"\band\b", " to ", time_text, flags=re.I)
    times = re.findall(r"\d{1,2}(?::\d{2})?\s*(?:AM|PM|NN|am|pm|nn)?|Morning|Afternoon|Evening", time_text)
    parsed = [parse_time(t) for t in times]
    parsed = [t for t in parsed if t]
    start_time = parsed[0] if parsed else "09:00"
    end_time = parsed[1] if len(parsed) > 1 else None
    if not end_time:
        start_dt = datetime.fromisoformat(f"{start_date.isoformat()}T{start_time}")
        end_time = (start_dt + timedelta(hours=1)).strftime("%H:%M")
    return start_date, end_date, start_time, end_time


def infer_layer(venue, subject, remarks, action):
    text = " ".join([venue, subject, remarks, action]).lower()
    if "reminder (deadline)" in venue.lower() or "deadline" in text or "submission" in text:
        return "Compliance"
    if any(word in text for word in ["training", "seminar", "orientation", "workshop", "briefing", "caucus"]):
        return "Training"
    if any(word in text for word in ["board", "bac", "committee", "management", "meeting", "conference"]):
        return "Management"
    if any(word in text for word in ["maintenance", "repair", "inspection", "substation", "lineworker", "crew"]):
        return "Maintenance"
    if any(word in text for word in ["project", "implementation", "procurement", "construction"]):
        return "Projects"
    if any(word in text for word in ["barangay", "assembly", "medical mission", "community"]):
        return "Enterprise-wide"
    return "Department" if any(code in text for code in ["nsd", "isd", "nnsd", "cpd", "pgd", "audit"]) else "Enterprise-wide"


def infer_department(*parts):
    text = " ".join(parts).lower()
    for code in ["ISD", "NNSD", "NSD", "CPD", "PGD", "AUD"]:
        if re.search(rf"\b{code.lower()}\b", text):
            return code
    if "audit" in text:
        return "AUD"
    return None


def event_uid(sheet, row, zimbra, title, start):
    key = zimbra.strip() or re.sub(r"[^A-Za-z0-9]+", "-", title[:48]).strip("-")
    return f"XLSX-{sheet}-{row}-{key}-{start}".replace(" ", "-")[:80]


def main():
    workbook_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    events = []
    skipped = []
    for ws in wb.worksheets:
        parsed_sheet = sheet_month_year(ws.title)
        if not parsed_sheet:
            continue
        sheet_month, sheet_year = parsed_sheet
        headers = [clean(ws.cell(1, c).value).upper() for c in range(1, min(ws.max_column, 8) + 1)]
        if "DATE AND TIME" not in headers or "SUBJECT" not in headers:
            continue
        for row in range(2, ws.max_row + 1):
            date_time = ws.cell(row, 1).value
            venue = clean(ws.cell(row, 2).value)
            subject = clean(ws.cell(row, 3).value)
            remarks = clean(ws.cell(row, 4).value)
            zimbra = clean(ws.cell(row, 5).value)
            action = clean(ws.cell(row, 6).value)
            if not clean(date_time) or not subject:
                continue
            parsed = date_range_and_time(date_time, sheet_month, sheet_year)
            if not parsed:
                skipped.append({"sheet": ws.title, "row": row, "dateTime": clean(date_time), "subject": subject})
                continue
            start_date, end_date, start_time, end_time = parsed
            title = subject[:300]
            description_bits = []
            if remarks:
                description_bits.append(f"Remarks: {remarks}")
            if action:
                description_bits.append(f"Action: {action}")
            if zimbra:
                description_bits.append(f"Zimbra Code: {zimbra}")
            description_bits.append(f"Source: {workbook_path.name} / {ws.title} row {row}")
            start = f"{start_date.isoformat()}T{start_time}:00"
            end = f"{end_date.isoformat()}T{end_time}:00"
            if end <= start:
                end_dt = datetime.fromisoformat(start) + timedelta(hours=1)
                end = end_dt.strftime("%Y-%m-%dT%H:%M:%S")
            events.append({
                "eventUid": event_uid(ws.title, row, zimbra, title, start),
                "title": title,
                "layer": infer_layer(venue, subject, remarks, action),
                "start": start,
                "end": end,
                "allDay": "T00:00:00" in start and "T00:00:00" in end,
                "location": venue if venue and not venue.upper().startswith("REMINDER") else None,
                "description": "\n\n".join(description_bits),
                "departmentCode": infer_department(venue, subject, remarks, action),
                "sourceRowKey": f"{ws.title}!{row}",
                "rawSource": json.dumps({
                    "dateTime": clean(date_time),
                    "venue": venue,
                    "subject": subject,
                    "remarks": remarks,
                    "zimbraCode": zimbra,
                    "action": action,
                }, ensure_ascii=False),
            })
    print(json.dumps({"events": events, "skipped": skipped}, ensure_ascii=False))


if __name__ == "__main__":
    main()
